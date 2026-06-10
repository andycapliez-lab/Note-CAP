import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration de la page Streamlit
st.set_page_config(page_title="Évaluation CAP Conducteur d'Engins", layout="centered")

st.title("🚜 Notation par Compétences - CAP")
st.write("Grille de notation officielle avec export Excel auto-ajusté.")

# --- 1. INFORMATIONS GÉNÉRALES ---
st.header("📋 Informations Générales")
col1, col2 = st.columns(2)

with col1:
    nom_candidat = st.text_input("Nom & Prénom du Candidat")
    formateur = st.text_input("Nom de l'Évaluateur")

with col2:
    date_eval = st.date_input("Date de l'évaluation", datetime.now())
    classe = st.selectbox(
        "Classe du Candidat",
        ["CAP 1 A", "CAP 1 B", "CAP 2 A", "CAP 2 B", "CAP 1AN", "Adulte / Formation Continue"]
    )
    engin = st.selectbox(
        "Type d'engin",
        ["Pelle Hydraulique", "Mini-pelle", "Chargeuse", "Tombereau", "Compacteur", "Bouteur", "Chariot"]
    )

st.divider()

# --- 2. GRILLE DE NOTATION ---
st.header("🎯 Barème par Compétences")

# --- SECTION 1 : PRISE DE POSTE ---
st.subheader("1. Prise de poste [C3.5] (/20)")
p1_visuel = st.slider("[C3.5] Contrôle visuel de l’engin", 0, 5, 5)
p1_niveaux = st.slider("[C3.5] Vérification des niveaux", 0, 5, 5)
p1_securite = st.slider("[C3.2] Équipements de sécurité (alarme, ceinture)", 0, 5, 5)
p1_docs = st.slider("[C2.1] Conformité documents (entretien, check-list)", 0, 5, 5)
total_p1 = p1_visuel + p1_niveaux + p1_securite + p1_docs
st.markdown(f"**Sous-total Section 1 : `{total_p1} / 20`**")

st.divider()

# --- SECTION 2 : CONDUITE ---
st.subheader("2. Conduite et Réalisation [C3.8 à C3.12] (/40)")
p2_commandes = st.slider("[C3.11] Maîtrise des commandes et fluidité", 0, 10, 7)
p2_securite = st.slider("[C3.2] Respect des consignes de sécurité et zones", 0, 10, 7)
p2_precision = st.slider("[C3.9/C3.11] Précision (terrassement, levage)", 0, 10, 7)
p2_terrain = st.slider("[C3.11] Adaptabilité aux contraintes terrain", 0, 10, 7)
total_p2 = p2_commandes + p2_securite + p2_precision + p2_terrain
st.markdown(f"**Sous-total Section 2 : `{total_p2} / 40`**")

st.divider()

# --- SECTION 3 : FIN DE POSTE ---
st.subheader("3. Fin de poste [C3.5] (/20)")
p3_nettoyage = st.slider("[C3.1] Nettoyage et rangement", 0, 5, 5)
p3_anomalies = st.slider("[C3.5] Rapport des anomalies ou pannes", 0, 5, 5)
p3_stationnement = st.slider("[C3.2] Stationnement en sécurité", 0, 5, 5)
p3_fiche = st.slider("[C3.5] Mise à jour de la fiche de poste", 0, 5, 5)
total_p3 = p3_nettoyage + p3_anomalies + p3_stationnement + p3_fiche
st.markdown(f"**Sous-total Section 3 : `{total_p3} / 20`**")

st.divider()

# --- SECTION 4 : COMPORTEMENT ---
st.subheader("4. Professionnalisme [C3.1] (/20)")
p4_consignes = st.slider("[C3.1] Respect des consignes orales/écrites", 0, 10, 8)
p4_attitude = st.slider("[C3.1] Attitude et communication pro", 0, 10, 8)
total_p4 = p4_consignes + p4_attitude
st.markdown(f"**Sous-total Section 4 : `{total_p4} / 20`**")

st.divider()

# --- 3. CALCUL DU RÉSULTAT FINAL SUR 20 ---
st.header("📊 Résultat Final")
note_totale_100 = total_p1 + total_p2 + total_p3 + total_p4
note_sur_20 = note_totale_100 / 5

st.metric(label="NOTE FINALE", value=f"{note_sur_20} / 20")

if note_sur_20 >= 10:
    st.success("🎉 **Validation réussie (Moyenne atteinte)**")
else:
    st.error("⚠️ **Insuffisant (En dessous de la moyenne)**")

commentaires = st.text_area("✍️ Observations générales (Points forts / Axes d'amélioration)")

st.divider()

# --- 4. EXPORT EXCEL ULTRA-STYLISÉ AVEC CASES AUTO-AJUSTÉES ---
if not nom_candidat:
    st.info("💡 Entrez le nom du candidat pour débloquer le bouton de téléchargement Excel.")
else:
    donnees_eval = {
        "Candidat": [nom_candidat],
        "Classe": [classe],
        "Évaluateur": [formateur],
        "Date": [str(date_eval)],
        "Engin": [engin],
        "C3.5_Visuel": [p1_visuel],
        "C3.5_Niveaux": [p1_niveaux],
        "C3.2_Secu_Poste": [p1_securite],
        "C2.1_Docs": [p1_docs],
        "TOTAL_P1": [total_p1],
        "C3.11_Commandes": [p2_commandes],
        "C3.2_Secu_Cond": [p2_securite],
        "C3.9_Precision": [p2_precision],
        "C3.11_Adapt": [p2_terrain],
        "TOTAL_P2": [total_p2],
        "C3.1_Nettoyage": [p3_nettoyage],
        "C3.5_Anomalies": [p3_anomalies],
        "C3.2_Stationnement": [p3_stationnement],
        "C3.5_Fiche": [p3_fiche],
        "TOTAL_P3": [total_p3],
        "C3.1_Consignes": [p4_consignes],
        "C3.1_Attitude": [p4_attitude],
        "TOTAL_P4": [total_p4],
        "NOTE BRUTE (/100)": [note_totale_100],
        "NOTE FINALE (/20)": [note_sur_20],
        "Observations": [commentaires]
    }
    
    df = pd.DataFrame(donnees_eval)
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Évaluation")
        worksheet = writer.sheets["Évaluation"]
        
        # Définition des styles graphiques (Comme l'ancien modèle)
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_body = Font(name="Arial", size=10)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Bleu marine professionnel
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        
        # 1. Application du style à l'en-tête (Ligne 1)
        worksheet.row_dimensions[1].height = 28
        for cell in worksheet[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # 2. Application du style aux valeurs (Ligne 2)
        worksheet.row_dimensions[2].height = 22
        for cell in worksheet[2]:
            cell.font = font_body
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", horizontal="left")
        
        # 3. Coloration intelligente de la note finale sur 20 (Colonne 25)
        cell_note = worksheet.cell(row=2, column=25)
        cell_note.font = Font(name="Arial", size=11, bold=True)
        if note_sur_20 >= 10:
            cell_note.fill = PatternFill(start_color="C6EFCE", fill_type="solid") # Vert
        else:
            cell_note.fill = PatternFill(start_color="FFC7CE", fill_type="solid") # Rouge

        # 4. LE RETOUR DE L'AJUSTEMENT AUTOMATIQUE DES COLONNES
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    # Permet de calculer la longueur réelle même s'il y a des retours à la ligne
                    lignes_texte = str(cell.value).split('\n')
                    for l in lignes_texte:
                        if len(l) > max_len:
                            max_len = len(l)
            
            col_letter = get_column_letter(col[0].column)
            # On donne une marge de +4 caractères. Taille minimale de 12, taille maximale de 45 (pour les longues observations)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
                
    buffer.seek(0)
    
    st.download_button(
        label="💾 Partager le Bilan de Compétences (Excel)",
        data=buffer,
        file_name=f"Eval_{classe.replace(' ', '_')}_{nom_candidat.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
